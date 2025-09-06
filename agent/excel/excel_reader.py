import re

import pandas as pd
from openpyxl import load_workbook


class ExcelReader:
    """
    Class for reading Excel files.
    """

    def __init__(self, file_name: str, file_path: str):
        """
        Initialize ExcelReader.

        Args:
            file_name (str): file name.
            file_path (str): path to file.
        """
        self.file_name = file_name
        self.file_path = file_path
        excel_file = pd.ExcelFile(file_path)
        self.sheets = excel_file.sheet_names
        self.dataframes = {}

    def _get_workbook(self):
        """Lazy loading workbook for cell operations."""
        return load_workbook(self.file_path, data_only=True)

    async def get_file_summary(self) -> dict:
        """
        Get summary of the file.

        Returns:
            dict: Summary of the file with keys: file_name, sheet_count, sheet_names
        """
        return {
            "file_name": self.file_name,
            "sheet_count": len(self.sheets),
            "sheet_names": self.sheets,
        }

    async def get_sheet_preview(self, sheet_name: str = None) -> dict:
        """
        Get preview of the specified sheet.

        Args:
            sheet_name (str): name of sheet for preview

        Returns:
            dict: Preview of the sheet with keys: sheet_name, row_count, column_count, columns, preview_rows, data_types
        """
        if sheet_name is None:
            sheet_name = self.sheets[0]

        if sheet_name not in self.dataframes:
            self.dataframes[sheet_name] = pd.read_excel(self.file_path, sheet_name)

        df = self.dataframes[sheet_name]

        return {
            "sheet_name": sheet_name,
            "row_count": len(df),
            "column_count": len(df.columns),
            "columns": df.columns.tolist(),
            "preview_rows": df.head(5).to_dict(orient="records"),
            "data_types": {col: str(dtype) for col, dtype in df.dtypes.items()},
        }

    async def search_data(self, sheet_name: str, search_term: str) -> dict:
        """
        Search data in the sheet by key word.

        Args:
            sheet_name (str): sheet name for search.
            search_term (str): search key word.

        Returns:
            dict: Search results with keys: sheet_name, search_term, matches_count, matches.
            Key error means wrong cell reference or out from bounds, so such cell is empty.
        """
        if sheet_name not in self.dataframes:
            self.dataframes[sheet_name] = pd.read_excel(self.file_path, sheet_name)

        df = self.dataframes[sheet_name]

        matches = []
        for col in df.columns:
            if df[col].dtype == "object":
                matches.extend(
                    df[
                        df[col]
                        .astype(str)
                        .str.contains(search_term, case=False, na=False)
                    ].to_dict(orient="records")
                )

        return {
            "sheet_name": sheet_name,
            "search_term": search_term,
            "matches_count": len(matches),
            "matches": matches[:10],  # Limit the number of results
        }

    async def get_cell_value(self, sheet_name: str, cell_reference: str) -> dict:
        """
        Get value of the cell by reference (A1, B2, etc.).

        Args:
            sheet_name (str): sheet name.
            cell_reference (str): reference of the cell.

        Returns:
            dict: Cell value with keys: sheet_name, cell_reference, value
        """
        try:
            wb = self._get_workbook()
            if sheet_name not in wb.sheetnames:
                return {"error": f"Sheet {sheet_name} not found"}

            ws = wb[sheet_name]

            if not re.match(r"^[A-Z]+\d+$", cell_reference.upper()):
                return {"error": f"Invalid cell reference format: {cell_reference}"}

            cell = ws[cell_reference.upper()]
            value = cell.value
            return {
                "sheet_name": sheet_name,
                "cell_reference": cell_reference.upper(),
                "value": value,
                "data_type": type(value).__name__ if value is not None else "NoneType"
            }

        except Exception as e:
            return {"error": f"Error accessing cell {cell_reference}: {str(e)}"}

    async def get_range_values(self, sheet_name: str, range_reference: str) -> dict:
        """
        Get values from a range of cells (e.g., 'A1:C3').

        Args:
            sheet_name (str): sheet name.
            range_reference (str): range reference (e.g., 'A1:C3').

        Returns:
            dict: Range values
        """
        try:
            wb = self._get_workbook()
            if sheet_name not in wb.sheetnames:
                return {"error": f"Sheet {sheet_name} not found"}

            ws = wb[sheet_name]

            cell_range = ws[range_reference]

            values = []
            if hasattr(cell_range, '__iter__') and not isinstance(cell_range, str):
                for row in cell_range:
                    if hasattr(row, '__iter__'):
                        row_values = []
                        for cell in row:
                            row_values.append({
                                "address": cell.coordinate,
                                "value": cell.value
                            })
                        values.append(row_values)
                    else:
                        values.append([{
                            "address": row.coordinate,
                            "value": row.value
                        }])
            else:
                values = [[{
                    "address": cell_range.coordinate,
                    "value": cell_range.value
                }]]

            return {
                "sheet_name": sheet_name,
                "range_reference": range_reference,
                "values": values
            }

        except Exception as e:
            return {"error": f"Error accessing range {range_reference}: {str(e)}"}

    async def analyze_column(self, sheet_name: str, column_name: str) -> dict:
        """
        Analyze the specified column.

        Args:
            sheet_name (str): sheet name.
            column_name (str): column name.

        Returns:
            dict: Analysis results with keys: sheet_name, column_name, stats
        """
        if sheet_name not in self.dataframes:
            self.dataframes[sheet_name] = pd.read_excel(self.file_path, sheet_name)

        df = self.dataframes[sheet_name]

        if column_name not in df.columns:
            return {"error": f"Column {column_name} not found"}

        column = df[column_name]

        # Basic statistics
        stats = {
            "count": len(column),
            "null_count": column.isna().sum(),
            "unique_values": column.nunique(),
        }

        # Additional statistics for numerical columns
        if pd.api.types.is_numeric_dtype(column):
            stats.update(
                {
                    "min": float(column.min()),
                    "max": float(column.max()),
                    "mean": float(column.mean()),
                    "median": float(column.median()),
                    "std": float(column.std()),
                }
            )

        # For categorical data
        else:
            # Most frequent values
            value_counts = column.value_counts().head(5).to_dict()
            stats["most_common_values"] = {
                str(k): int(v) for k, v in value_counts.items()
            }

        return {"sheet_name": sheet_name, "column_name": column_name, "stats": stats}
